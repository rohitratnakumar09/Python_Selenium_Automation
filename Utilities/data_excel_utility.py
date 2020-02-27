""" This module is used for developing/ accessing data reader utility. """

import os
import logging

import openpyxl
from Utilities.config_utility import config_utility
import Utilities.logger_utility as log_utils
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
class DataReader:
    """
    This class includes basic reusable data helpers.
    """

    log = log_utils.custom_logger(logging.INFO)
    config = config_utility()
    prop = config.load_config_file()
    base_test_data = prop.get('PROD', 'test_data')
    directory = os.path.dirname(os.getcwd())
    def __init__(self):
        self.ui_file_path = os.path.join(self.directory, self.base_test_data)


    def load_excel(self):
        book =None


        try:
            if self.ui_file_path is not None:
                book = openpyxl.load_workbook(self.ui_file_path)
        except Exception as ex:
            self.log.error("Failed to load test data.", ex)
        return book
    def get_excel_data(self,sheet_name,tc_name):
        value = {}
        book = self.load_excel()
        sheet = book[sheet_name]
        try:
            if book is not None:
                for i in range(1, sheet.max_row + 1):  # to get rows
                    if sheet.cell(row=i, column=1).value == tc_name:
                        for j in range(2, sheet.max_column + 1):  # to get columns
                            value[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        except Exception as ex:
            self.log.error("Failed to load test data.", ex)
        return value

    def add_excel_result(self,sheet_name,tc_name,result):

        book = self.load_excel()
        sheet = book[sheet_name]
        try:
            if book is not None:

                for i in range(1, sheet.max_row + 1):  # to get rows
                    if sheet.cell(row=i, column=1).value == tc_name:
                        for j in range(2, sheet.max_column + 1):
                            if sheet.cell(row=1, column=j).value=="Result":
                                if result=="Pass":
                                    sheet.cell(row=i, column=j).value="Pass"
                                    green=openpyxl.styles.colors.Color(rgb='0000ff00')
                                    fill_cell = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=green)
                                    sheet.cell(row=i, column=j).fill = fill_cell
                                else:
                                    sheet.cell(row=i, column=j).value = "Fail"
                                    red = openpyxl.styles.colors.Color(rgb='00FF0000')
                                    fill_cell = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=red)
                                    sheet.cell(row=i, column=j).fill = fill_cell
                book.save(self.ui_file_path)
        except Exception as ex:
            self.log.error("Failed to add result to test data.", ex)
